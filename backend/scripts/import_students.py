"""
원생목록.xlsx에서 학생 데이터를 읽어 일괄 등록하는 스크립트

Usage:
    cd backend
    python -m scripts.import_students                    # 드라이런 (미리보기)
    python -m scripts.import_students --run              # 실제 등록
    python -m scripts.import_students --run --teacher 이메일  # 특정 교사에게 배정

엑셀 매핑:
    - username = 이름 (col 0)
    - password = 출결번호 (col 2), 없으면 "0000"
    - name     = 이름 (col 0)
    - school   = 학교 (col 3)
    - grade    = 학년 (col 4)
"""
import asyncio
import argparse
import sys
from pathlib import Path

# Windows 콘솔 UTF-8 출력
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from sqlalchemy import select, func

# ── 프로젝트 루트를 sys.path에 추가 ──
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db.session import get_db
from app.models.user import User
from app.core.security import get_password_hash

XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "원생목록.xlsx"

# 엑셀 컬럼 인덱스
COL_NAME = 0
COL_GENDER = 1
COL_ATTENDANCE_NO = 2   # 출결번호 → password
COL_SCHOOL = 3
COL_GRADE = 4


def read_xlsx() -> list[dict]:
    """엑셀 파일에서 학생 데이터를 읽어온다."""
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb[wb.sheetnames[0]]

    students = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        name = (row[COL_NAME] or "").strip()
        if not name:
            continue

        att_no = row[COL_ATTENDANCE_NO]
        # 출결번호: 숫자/문자열 모두 처리, 없으면 "0000"
        if att_no is not None and str(att_no).strip():
            password = str(int(att_no)) if isinstance(att_no, float) else str(att_no).strip()
        else:
            password = "0000"

        school = (row[COL_SCHOOL] or "").strip() or None
        grade = (row[COL_GRADE] or "").strip() or None

        students.append({
            "name": name,
            "username": name,       # 이름을 username으로 사용
            "password": password,
            "school_name": school,
            "grade": grade,
        })

    wb.close()
    return students


async def get_teacher(db, teacher_email: str | None) -> User:
    """교사 계정을 찾는다. 지정 안 하면 첫 번째 교사 사용."""
    if teacher_email:
        result = await db.execute(
            select(User).where(User.role == "teacher", User.email == teacher_email)
        )
        teacher = result.scalar_one_or_none()
        if not teacher:
            print(f"[ERROR] 교사 '{teacher_email}'을 찾을 수 없습니다.")
            sys.exit(1)
        return teacher

    # 기본: 첫 번째 교사
    result = await db.execute(
        select(User).where(User.role == "teacher").order_by(User.created_at).limit(1)
    )
    teacher = result.scalar_one_or_none()
    if not teacher:
        print("[ERROR] 등록된 교사가 없습니다. 먼저 교사 계정을 만드세요.")
        sys.exit(1)
    return teacher


async def main():
    parser = argparse.ArgumentParser(description="원생목록.xlsx → 학생 일괄 등록")
    parser.add_argument("--run", action="store_true", help="실제 DB에 등록 (없으면 드라이런)")
    parser.add_argument("--teacher", type=str, default=None, help="교사 이메일 (기본: 첫 번째 교사)")
    args = parser.parse_args()

    # 1) 엑셀 읽기
    students = read_xlsx()
    print(f"엑셀 파일: {XLSX_PATH}")
    print(f"읽은 학생 수: {len(students)}명\n")

    if not students:
        print("[WARN] 등록할 학생이 없습니다.")
        return

    # 2) DB 세션
    async for db in get_db():
        teacher = await get_teacher(db, args.teacher)
        print(f"담당 교사: {teacher.name} ({teacher.email or teacher.id[:8]})")

        # 기존 학생 조회 (해당 교사 소속)
        existing_result = await db.execute(
            select(User.username).where(
                User.role == "student",
                User.teacher_id == teacher.id,
            )
        )
        existing_usernames = {r[0] for r in existing_result.all()}

        # 전체 username 조회 (글로벌 중복 체크)
        global_result = await db.execute(select(User.username).where(User.username.isnot(None)))
        global_usernames = {r[0] for r in global_result.all()}

        # 3) 분류
        to_create = []
        skipped_dup = []
        skipped_global = []

        seen_in_batch = set()
        for s in students:
            uname = s["username"]

            # 배치 내 중복 (동명이인)
            if uname in seen_in_batch:
                skipped_dup.append(s)
                continue
            seen_in_batch.add(uname)

            # 이미 같은 교사에 등록됨
            if uname in existing_usernames:
                skipped_dup.append(s)
                continue

            # 글로벌 username 충돌
            if uname in global_usernames:
                skipped_global.append(s)
                continue

            to_create.append(s)

        # 4) 리포트
        print(f"\n{'='*50}")
        print(f"  신규 등록 예정: {len(to_create)}명")
        print(f"  이미 등록 (스킵): {len(skipped_dup)}명")
        if skipped_global:
            print(f"  글로벌 username 충돌: {len(skipped_global)}명")
        print(f"{'='*50}")

        if skipped_dup:
            print(f"\n[스킵 - 이미 등록] {len(skipped_dup)}명:")
            for s in skipped_dup[:10]:
                print(f"  - {s['name']} ({s['school_name']} {s['grade']})")
            if len(skipped_dup) > 10:
                print(f"  ... 외 {len(skipped_dup) - 10}명")

        if skipped_global:
            print(f"\n[스킵 - username 충돌] {len(skipped_global)}명:")
            for s in skipped_global:
                print(f"  - {s['name']} (username '{s['username']}' 이미 사용 중)")

        if to_create:
            print(f"\n[등록 예정] {len(to_create)}명 (처음 10명 미리보기):")
            for s in to_create[:10]:
                pw_display = s["password"] if s["password"] != "0000" else "0000 (출결번호 없음)"
                print(f"  - {s['name']} | 학교={s['school_name']} | 학년={s['grade']} | pw={pw_display}")
            if len(to_create) > 10:
                print(f"  ... 외 {len(to_create) - 10}명")

        # 5) 드라이런이면 여기서 종료
        if not args.run:
            print(f"\n[드라이런] --run 옵션을 추가하면 실제 등록됩니다.")
            return

        # 6) 실제 등록
        print(f"\n등록 시작...")
        created = 0
        errors = 0
        for i, s in enumerate(to_create):
            try:
                student = User(
                    username=s["username"],
                    password_hash=get_password_hash(s["password"]),
                    name=s["name"],
                    role="student",
                    teacher_id=teacher.id,
                    school_name=s["school_name"],
                    grade=s["grade"],
                )
                db.add(student)
                created += 1

                # 50명마다 중간 커밋 + 진행률 표시
                if created % 50 == 0:
                    await db.flush()
                    print(f"  ... {created}/{len(to_create)} 완료")
            except Exception as e:
                errors += 1
                print(f"  [ERROR] {s['name']}: {e}")

        await db.commit()
        print(f"\n{'='*50}")
        print(f"  등록 완료: {created}명")
        if errors:
            print(f"  오류: {errors}명")
        print(f"{'='*50}")

        # 최종 확인
        count_result = await db.execute(
            select(func.count()).where(
                User.role == "student",
                User.teacher_id == teacher.id,
            )
        )
        total = count_result.scalar()
        print(f"\n{teacher.name} 교사 총 학생 수: {total}명")


if __name__ == "__main__":
    asyncio.run(main())
